import os
from dotenv import load_dotenv
from pathlib import Path
from openpyxl import load_workbook
from langsmith import Client

# Load .env from project root
load_dotenv(dotenv_path=Path(__file__).parent.parent / ".env")

# Initialize LangSmith client
client = Client()

# Check if dataset already exists
dataset_name = "dena-ime-portfolio-reports"
existing_datasets = [ds.name for ds in client.list_datasets()]
if dataset_name in existing_datasets:
    print("Dataset already exists, skipping.")
    exit()

# Read Excel file
excel_path = Path(__file__).parent.parent / "data" / "dena_IME_mock_dataset_v2.xlsx"
wb = load_workbook(excel_path, data_only=True)

# Helper function to extract data from sheets
def get_sheet_data(sheet_name, start_row=3):
    """Extract data from a sheet, starting from start_row."""
    ws = wb[sheet_name]
    data = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx < start_row:
            continue
        if not any(cell for cell in row):  # Skip empty rows
            break
        data.append(row)
    return data

# Load data from sheets
consolidated = get_sheet_data("CONSOLIDATED DASHBOARD", start_row=4)
sap_finance = get_sheet_data("SOURCE 1 — SAP Finance", start_row=4)
pl_status = get_sheet_data("SOURCE 2 — PL Status Input", start_row=4)
milestones = get_sheet_data("SOURCE 4 — Milestone Tracker", start_row=4)

# Parse consolidated dashboard
consolidated_by_id = {}
for row in consolidated:
    if row[0]:  # Project ID
        project_id = str(row[0]).strip()
        consolidated_by_id[project_id] = {
            "project_name": row[1],
            "status": row[8] if len(row) > 8 else None,  # Overall Status
            "risk_flag": row[9] if len(row) > 9 else None,  # Escalation?
            "next_milestone": row[10] if len(row) > 10 else None,
        }

# Parse SAP Finance
finance_by_id = {}
for row in sap_finance:
    if row[0]:  # Project ID
        project_id = str(row[0]).strip()
        finance_by_id[project_id] = {
            "budget_total": float(row[3]) if row[3] else 0,  # Contract Value
            "budget_spent_ytd": float(row[6]) if row[6] else 0,  # Actual Spend YTD
        }

# Parse PL Status
status_by_id = {}
for row in pl_status:
    if row[0]:  # Project ID
        project_id = str(row[0]).strip()
        status_by_id[project_id] = {
            "project_lead": row[2] if len(row) > 2 else None,
            "monthly_status_text": row[7] if len(row) > 7 else None,  # Free Text Update
        }

# Parse Milestones
milestones_by_id = {}
for row in milestones:
    if row[0]:  # Project ID
        project_id = str(row[0]).strip()
        if project_id not in milestones_by_id:
            milestones_by_id[project_id] = []
        milestones_by_id[project_id].append({
            "milestone_name": row[2] if len(row) > 2 else None,
            "milestone_due_date": row[3] if len(row) > 3 else None,
            "milestone_status": row[5] if len(row) > 5 else None,
        })

# Join data and create examples
examples = []
project_ids = list(set(consolidated_by_id.keys()) & set(finance_by_id.keys()) & set(status_by_id.keys()))
project_ids = sorted(project_ids)[:12]  # Take first 12

for project_id in project_ids:
    consolidated_data = consolidated_by_id.get(project_id, {})
    finance_data = finance_by_id.get(project_id, {})
    status_data = status_by_id.get(project_id, {})
    milestone_list = milestones_by_id.get(project_id, [{}])
    next_milestone = milestone_list[0] if milestone_list else {}

    example = {
        "inputs": {
            "project_id": project_id,
            "project_name": consolidated_data.get("project_name"),
            "status": consolidated_data.get("status"),
            "risk_flag": consolidated_data.get("risk_flag"),
            "project_lead": status_data.get("project_lead"),
            "budget_total": finance_data.get("budget_total", 0),
            "budget_spent_ytd": finance_data.get("budget_spent_ytd", 0),
            "monthly_status_text": status_data.get("monthly_status_text"),
            "next_milestone": next_milestone.get("milestone_name"),
            "next_milestone_due": next_milestone.get("milestone_due_date"),
            "next_milestone_status": next_milestone.get("milestone_status"),
        },
        "outputs": {
            "expected_summary_contains": [
                "project name",
                "risk flag or status",
                "budget utilization",
                "next milestone",
                "project lead name",
            ]
        },
    }
    examples.append(example)

# Create dataset
dataset = client.create_dataset(
    dataset_name=dataset_name,
    description="LangSmith evaluation dataset for dena IME portfolio report generation"
)

# Upload examples
for example in examples:
    client.create_example(
        inputs=example["inputs"],
        outputs=example["outputs"],
        dataset_id=dataset.id,
    )

print(f"✓ Dataset '{dataset_name}' created successfully.")
print(f"✓ Uploaded {len(examples)} examples.")
