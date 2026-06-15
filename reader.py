import logging
from pathlib import Path
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def read_excel_data(file_path: str) -> list[dict]:
    """
    Read project portfolio data from Excel file.

    Args:
        file_path: Path to .xlsx file

    Returns:
        List of dicts with project data, or empty list if file not found
    """
    path = Path(file_path)

    if not path.exists():
        logger.error(f"Data file not found: {file_path}")
        return []

    try:
        wb = load_workbook(path)
        ws = wb["CONSOLIDATED DASHBOARD"]

        data = []
        column_headers = [
            "project_id",
            "project_name",
            "funder",
            "budget",
            "spend_ytd",
            "budget_pct",
            "hr_utilisation",
            "milestone_status",
            "overall_status",
            "escalation",
            "next_milestone",
            "key_risk"
        ]

        # Skip first 3 rows (title, subtitle, header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_col=12, values_only=True), start=4):
            # Stop if column A (project_id) is empty
            if row[0] is None:
                break

            # Create dict from row data
            project_dict = {}
            for col_idx, header in enumerate(column_headers):
                project_dict[header] = row[col_idx]

            data.append(project_dict)

        logger.info(f"Successfully read {len(data)} projects from {file_path}")
        return data

    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        return []
